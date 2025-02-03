from django.shortcuts import render , redirect
from . forms import CreateRecord , VisitorEditForm
from . models import Coordinator , Visitor 
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import authenticate , login , logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Q
from datetime import date
from django.core.paginator import Paginator
import openpyxl
from django.http import HttpResponse


@login_required(login_url='login')
def home(request,q=None):

    if request.method == "POST":
        name = request.POST.get('name')
        oragnization = request.POST.get('organization')
        location = request.POST.get('location')
        domain = request.POST.get('domain')

    #     visit = Visitor.objects.filter(
    #     Q(name__icontains = name) &
    #     Q(location__icontains = location) &
    #     Q(domain__icontains = domain) &
    #     Q(organization__icontains = oragnization) 
    # )
        
        visit = Visitor.objects.filter(
        name__icontains = name,
        location__icontains = location,
        domain__icontains = domain,
        organization__icontains = oragnization)

        p = Paginator(visit , 2)
        page = request.GET.get('page')
        pages = p.get_page(page)
        
    else:
        q = request.GET.get('q') if request.GET.get('q') != None else ''

        visit = Visitor.objects.filter(
            Q(name__icontains = q) |
            Q(location__icontains = q) |
            Q(domain__icontains = q) |
            Q(organization__icontains = q) |
            Q(event__icontains = q) |
            Q(host__name__icontains = q)
        )

        p = Paginator(visit , 2)
        page = request.GET.get('page')
        pages = p.get_page(page)
    
    year1 = Visitor.objects.filter(
        Q(date__gte = date(year=2021, month=4, day=1) , date__lte = date(year=2022, month=3, day=31))
    )
    year2 = Visitor.objects.filter(
        Q(date__gte = date(year=2022, month=4, day=1) , date__lte = date(year=2023, month=3, day=31))
    )
    year3 = Visitor.objects.filter(
        Q(date__gte = date(year=2023, month=4, day=1) , date__lte = date(year=2024, month=3, day=31))
    )
    year4 = Visitor.objects.filter(
        Q(date__gte = date(year=2024, month=4, day=1) , date__lte = date(year=2025, month=3, day=31))
    )

    count = Visitor.objects.all().count()

    context = {'visitors':visit , 'year1':year1, 'year2':year2, 'year3':year3, 'year4':year4 , 'pages':pages , 'count':count }
    return render(request, 'app1/home.html' , context)



def year_fil(request ,ys,ye):
    # ys = request.GET.get('ys')
    # ye = request.GET.get('ye')

    financial_year_start = date(year=ys, month=4, day=1)
    financial_year_end = date(year=ye, month=3, day=31)

    count = Visitor.objects.all().count()
    
    visit = Visitor.objects.filter(
        Q(date__gte = financial_year_start , date__lte = financial_year_end)
    )
    p = Paginator(visit , 2)
    page = request.GET.get('page')
    pages = p.get_page(page)

    year1 = Visitor.objects.filter(
        Q(date__gte = date(year=2021, month=4, day=1) , date__lte = date(year=2022, month=3, day=31))
    )
    year2 = Visitor.objects.filter(
        Q(date__gte = date(year=2022, month=4, day=1) , date__lte = date(year=2023, month=3, day=31))
    )
    year3 = Visitor.objects.filter(
        Q(date__gte = date(year=2023, month=4, day=1) , date__lte = date(year=2024, month=3, day=31))
    )
    year4 = Visitor.objects.filter(
        Q(date__gte = date(year=2024, month=4, day=1) , date__lte = date(year=2025, month=3, day=31))
    )

    context = {'visitors':visit , 'year1':year1, 'year2':year2, 'year3':year3, 'year4':year4 , 'pages':pages , 'count':count }
    return render(request, 'app1/home.html' , context)



@login_required(login_url='login')
def create(request):

    form = CreateRecord()

    if request.method == 'POST':
        form = CreateRecord(request.POST)

        if form.is_valid():
            coordinator_name = form.cleaned_data['coordinator_name']
            coordinator_phone = form.cleaned_data['coordinator_phone']
            visitor_name = form.cleaned_data['visitor_name']
            phone = form.cleaned_data['phone']
            organization = form.cleaned_data['organization']
            location = form.cleaned_data['location']
            domain = form.cleaned_data['domain']
            event = form.cleaned_data['event']
            date = form.cleaned_data['date']

            try:
                coordinator  = Coordinator.objects.get(name = coordinator_name , phone = coordinator_phone)

            except Coordinator.DoesNotExist:

                coordinator  = Coordinator.objects.create(name = coordinator_name , phone = coordinator_phone)
                coordinator.save()
            
            visitor = Visitor.objects.create(
                host = coordinator,
                name = visitor_name,
                organization = organization,
                phone=phone,
                location = location,
                domain = domain,
                event = event,
                date = date
            )
            visitor.save()

            return redirect('home')

    context = {'form':form}
    return render(request , 'app1/create.html' ,context)



@login_required(login_url='login')
def edit(request ,pk):

    visit = Visitor.objects.get(id=pk)
    form = VisitorEditForm(instance = visit)
    if request.method == "POST":
        form = VisitorEditForm(request.POST , instance = visit)
        if form.is_valid():
            form.save()
            return redirect('home')
    
    context = {'form':form}
    return render(request , 'app1/edit.html' ,context)


@login_required(login_url='login')
def delete(request , pk):
    visit = Visitor.objects.get(id=pk)

    if request.method == "POST":
        visit.delete()
        return redirect('home')

    context = {'obj':visit}
    return render(request , 'app1/delete.html' , context)


def loginUser(request):
    page = 'login'

    if request.method =="POST":
        username = request.POST.get('username').lower()
        password = request.POST.get('password')

        user = authenticate(request , username=username , password=password)

        if user is not None:
            login(request,user)
            return redirect('home')


    context = {'page':page}
    return render(request , 'app1/login.html' , context)



def register(request):

    form = UserCreationForm()

    if request.method == "POST":
        form = UserCreationForm(request.POST)

        if form.is_valid():
            user = form.save(commit=False)
            user.username = user.username.lower()
            user.save()
            login(request , user)
            return redirect('home')


    context = {'form':form}
    return render(request , 'app1/login.html' , context)



def logoutUser(request):
    logout(request)
    return redirect('home')


def export_to_excel(request):
    page_number = request.GET.get('page', 1)
    

    visitors = Visitor.objects.all()  
    paginator = Paginator(visitors, 10)

    current_page_visitors = paginator.get_page(page_number)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Visitor Details'

    columns = ['Name', 'Organization', 'Location', 'Domain', 'Phone', 'Event', 'Date', 'Coordinator Name', 'Coordinator Phone']

    for col_num, column_title in enumerate(columns, 1):
        worksheet.cell(row=1, column=col_num, value=column_title)

    for row_num, visitor in enumerate(current_page_visitors, 2):
        row_data = [
            visitor.name,
            visitor.organization,
            visitor.location,
            visitor.domain,
            visitor.phone,
            visitor.event,
            visitor.date,
            visitor.host.name if visitor.host else '',
            visitor.host.phone if visitor.host else ''
        ]

        for col_num, cell_value in enumerate(row_data, 1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=visitors_page_{}.xlsx'.format(page_number)
    
    workbook.save(response)

    return response

